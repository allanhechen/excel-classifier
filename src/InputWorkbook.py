from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from Expense import Expense


EXPECTED_HEADERS = ["Title", "Description", "Debit", "Credit", "ChargePostedDate"]


class InputWorkbook:
    @staticmethod
    def read_workbook(file_path: str) -> set[Expense]:
        workbook = load_workbook(file_path, data_only=True)
        result: set[Expense] = set()

        for sheet_name in workbook.sheetnames:
            sheet_expenses = InputWorkbook._read_worksheet(workbook[sheet_name])
            result = result.union(sheet_expenses)

        return result

    @staticmethod
    def _read_worksheet(worksheet: Worksheet) -> set[Expense]:
        result: set[Expense] = set()
        headers = [
            cell if cell is not None else ""
            for cell in next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )
        ]

        if headers != EXPECTED_HEADERS:
            raise ValueError(
                f"Unexpected headers: {headers}. Expected: {EXPECTED_HEADERS}"
            )

        objects = [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]

        for object in objects:
            if type(object["Title"]) is not str:
                raise Exception
            if type(object["Description"]) is not str:
                object["Description"] = ""
            if type(object["Debit"]) is not float:
                object["Debit"] = 0.0
            if type(object["Credit"]) is not float:
                object["Credit"] = 0.0
            if type(object["ChargePostedDate"]) is not int:
                raise Exception

            result.add(
                Expense(
                    object["Title"],
                    object["Description"],
                    worksheet.title,
                    int(float(object["Debit"]) * 100),
                    int(float(object["Credit"]) * 100),
                    InputWorkbook._excel_date_to_python_date(
                        object["ChargePostedDate"]
                    ),
                )
            )

        return result

    @staticmethod
    def _excel_date_to_python_date(excel_date) -> date:
        excel_base_date = date(1900, 1, 1)

        if excel_date < 61:
            return excel_base_date + timedelta(days=excel_date - 2)
        else:
            return excel_base_date + timedelta(days=excel_date - 1)
